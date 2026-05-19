import json
import re
import os
import sys
import unicodedata
from datetime import datetime
from openpyxl import load_workbook

# Forzar UTF-8 en la salida
if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8')

# Tabla de reemplazos para caracteres mal codificados
REEMPLAZOS = {
    # Vocales con acento (Windows-1252 a UTF-8)
    'Ã¡': 'á', 'Ã©': 'é', 'Ã­': 'í', 'Ã³': 'ó', 'Ãº': 'ú',
    'Ã€': 'À', 'Ãˆ': 'È', 'ÃŒ': 'Ì', 'Ã’: 'Ò', 'Ã': 'Ù',
    'Ã¢': 'â', 'Ãª': 'ê', 'Ã®': 'î', 'Ã´': 'ô', 'Ã»': 'û',
    # Ñ y Ü
    'Ã±': 'ñ', 'Ã‘': 'Ñ', 'Ã¼': 'ü', 'Ãœ': 'Ü',
    # Símbolos comunes
    'Â¡': '¡', 'Â¿': '¿', 'Â°': '°', 'Â·': '·', 'Â®': '®',
    'Â©': '©', 'Â§': '§', 'Â¶': '¶',
    # Comillas y apóstrofes
    'â€œ': '"', 'â€': '"', 'â€˜': "'", 'â€™': "'",
    'â€š': ',', 'â€ž': '"', 'â€¦': '...',
    # Guiones
    'â€”': '-', 'â€“': '-',
    
    # Problemas específicos de laboratorios
    'Bag¿': 'Bagó',
    'bag¿': 'Bagó',
    'Cassarß': 'Cassará',
    'cassarß': 'Cassará',
    'Temis-Lostal¿': 'Temis-Lostaló',
    'G¿minis Farmac¿': 'Géminis Farmacéutica',
    'Andr¿maco': 'Andrómaco',
    
    # Reemplazos generales para caracteres sueltos
    '¿': 'ó',
    'ß': 'á',
    
    # Casos con espacios
    '¿ ': 'ó ',
    'ß ': 'á ',
}

def normalizar_nombre(texto):
    """Convierte a formato título: primeras letras mayúsculas"""
    if not texto:
        return ''
    # Excepciones que no se capitalizan
    excepciones = {'y', 'de', 'la', 'del', 'los', 'las', 'con', 'sin', 'por', 'para', 'a', 'ante', 'bajo', 'cabe', 'contra', 'desde', 'durante', 'en', 'entre', 'hacia', 'hasta', 'mediante', 'segun', 'so', 'sobre', 'tras', 'el', 'un', 'una', 'y/o', 'e', 'u'}
    
    palabras = texto.split()
    resultado = []
    for i, p in enumerate(palabras):
        if i == 0 or p.lower() not in excepciones:
            resultado.append(p[0].upper() + p[1:].lower() if len(p) > 1 else p.upper())
        else:
            resultado.append(p.lower())
    return ' '.join(resultado)

def limpiar_texto(texto):
    """Limpia y normaliza texto, corrige acentos mal codificados"""
    if not texto:
        return ''
    texto = str(texto)
    
    # Reemplazar caracteres mal codificados
    for mal, bien in REEMPLAZOS.items():
        texto = texto.replace(mal, bien)
    
    # Normalizar Unicode
    texto = unicodedata.normalize('NFC', texto)
    
    # Eliminar caracteres no imprimibles
    texto = re.sub(r'[^\x20-\x7E\xA0-\xFF\u0100-\uFFFF]', '', texto)
    
    # Eliminar espacios múltiples
    texto = re.sub(r'\s+', ' ', texto).strip()
    
    # Normalizar formato del nombre
    texto = normalizar_nombre(texto)
    
    return texto

def limpiar_precio(valor):
    """Limpia el precio y lo convierte a float"""
    if not valor:
        return 0
    precio_raw = str(valor)
    precio_limpio = re.sub(r'[^\d.,-]', '', precio_raw).replace(',', '.')
    try:
        return float(precio_limpio)
    except:
        return 0

# Buscar el archivo XLSX
archivo_xls = None
for archivo in os.listdir('.'):
    if archivo.endswith('.xlsx'):
        archivo_xls = archivo
        break

if not archivo_xls:
    print("❌ No hay archivo .xlsx en esta carpeta")
    exit(1)

print(f"📄 Leyendo: {archivo_xls}")
wb = load_workbook(archivo_xls, data_only=True)
ws = wb.active

medicamentos = []

for row in ws.iter_rows(min_row=2, values_only=True):
    if not row[0]:
        continue
    
    droga = limpiar_texto(row[0])
    marca = limpiar_texto(row[1])
    presentacion = limpiar_texto(row[2])
    laboratorio = limpiar_texto(row[3])
    cobertura = str(row[4]).replace('%', '').strip() if row[4] else '0'
    copago = limpiar_precio(row[5])
    
    medicamentos.append({
        "DROGA": droga,
        "MARCA": marca,
        "PRESENTACION": presentacion,
        "LABORATORIO": laboratorio,
        "COBERTURA": cobertura,
        "COPAGO": copago
    })

fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M")
datos = {"fecha": fecha_actual, "medicamentos": medicamentos}

with open('medicamentos.json', 'w', encoding='utf-8') as f:
    json.dump(datos, f, ensure_ascii=False, indent=2)

print(f"✅ {len(medicamentos)} medicamentos guardados")
print(f"📅 Fecha: {fecha_actual}")

# Mostrar laboratorios para verificar
labs = {}
for med in medicamentos:
    lab = med['LABORATORIO']
    if lab and lab not in labs:
        labs[lab] = True
print("\n🔤 Muestra de laboratorios:")
for i, lab in enumerate(sorted(labs.keys())[:20]):
    print(f"   {i+1}. {lab}")